import os
import datetime
from flask import Flask, render_template_string, request, redirect, url_for, flash
import openpyxl

# --- Configuration ---
# This script creates a complete Flask application in a single file.
# It will generate the necessary HTML and CSS files automatically.

# Define the absolute base directory of the script to ensure paths are correct
# regardless of where the script is run from.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define the folder structure using absolute paths
STATIC_FOLDER = os.path.join(BASE_DIR, 'static')
CSS_FILE = os.path.join(STATIC_FOLDER, 'style.css')
EXCEL_FILE = os.path.join(BASE_DIR, 'weights.xlsx')

# --- HTML Content ---
# This is the HTML for our web page.
HTML_CONTENT = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Weight & Body Tracker</title>
    <link rel="stylesheet" href="{{ url_for('static', filename='style.css') }}">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3.0.1/dist/chartjs-plugin-annotation.min.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
</head>
<body>
    <div class="container">
        <h1>Weight & Body Tracker 🏋️</h1>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                <div class="flash-messages">
                {% for category, message in messages %}
                    <div class="flash {{ category }}">{{ message }}</div>
                {% endfor %}
                </div>
            {% endif %}
        {% endwith %}

        <div class="main-grid">
            <div class="left-column">
                <div class="card">
                    <h2>User Selection & Management</h2>
                    <div class="user-management-grid">
                        <form action="/" method="get" class="user-selector-form">
                            <div class="form-group">
                                <label for="user1">Primary User:</label>
                                <select name="user1" id="user1" onchange="this.form.submit()">
                                    {% for u in all_users %}
                                    <option value="{{ u }}" {% if u == primary_user %}selected{% endif %}>{{ u }}</option>
                                    {% endfor %}
                                </select>
                            </div>
                            <div class="form-group">
                                <label for="user2">Compare With (Optional):</label>
                                <select name="user2" id="user2" onchange="this.form.submit()">
                                    <option value="">-- None --</option>
                                    {% for u in all_users %}
                                        {% if u != primary_user %}
                                        <option value="{{ u }}" {% if u == comparison_user %}selected{% endif %}>{{ u }}</option>
                                        {% endif %}
                                    {% endfor %}
                                </select>
                            </div>
                        </form>
                        <div class="user-actions">
                            <form action="{{ url_for('add_user') }}" method="post" class="add-user-form">
                                <label for="new_user_name">Add New User:</label>
                                <div class="input-with-button">
                                    <input type="text" name="new_user_name" id="new_user_name" placeholder="Enter name" required>
                                    <button type="submit" class="btn btn-primary">Add</button>
                                </div>
                            </form>
                            <form action="{{ url_for('delete_user') }}" method="post">
                                <input type="hidden" name="user" value="{{ primary_user }}">
                                <button type="submit" class="btn btn-danger" onclick="return confirm('Are you sure you want to delete {{ primary_user }} and all their data? This cannot be undone.');">Delete Primary User</button>
                            </form>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <h2>Log New Entry for {{ primary_user }}</h2>
                    <form action="{{ url_for('index') }}" method="post" class="log-entry-form">
                        <input type="hidden" name="user" value="{{ primary_user }}">
                        <div class="log-entry-form-grid">
                            <div class="form-group">
                                <label for="weight">Current Weight (lbs):</label>
                                <input type="number" id="weight" name="weight" step="0.01" required placeholder="e.g., 175.50">
                            </div>
                            <div class="form-group">
                                <label for="date">Date:</label>
                                <input type="date" id="date" name="date" value="{{ today_date }}" required>
                            </div>
                            <div class="form-group">
                                <label for="body_fat">Body Fat (%): <span class="optional-text">(Optional)</span></label>
                                <input type="number" id="body_fat" name="body_fat" step="0.01" placeholder="e.g., 21.50">
                            </div>
                            <div class="form-group">
                                <label for="waist_size">Waist Size (in): <span class="optional-text">(Optional)</span></label>
                                <input type="number" id="waist_size" name="waist_size" step="0.01" placeholder="e.g., 34.55">
                            </div>
                        </div>
                        <button type="submit" class="btn btn-primary">Add Entry</button>
                    </form>
                </div>
                <div class="card">
                    <h2>{{ primary_user }}'s Goals</h2>
                    <form action="{{ url_for('update_goals') }}" method="post" class="goal-form">
                        <input type="hidden" name="user" value="{{ primary_user }}">
                        <div class="form-group">
                            <label for="start_weight">Start Weight (lbs):</label>
                            <input type="number" id="start_weight" name="start_weight" step="0.01" placeholder="e.g., 180.00" value="{{ primary_user_data.start_weight or '' }}">
                        </div>
                        <div class="form-group">
                            <label for="goal_weight">Goal Weight (lbs):</label>
                            <input type="number" id="goal_weight" name="goal_weight" step="0.01" placeholder="e.g., 165.00" value="{{ primary_user_data.goal_weight or '' }}">
                        </div>
                        <button type="submit" class="btn btn-secondary">Save Goals</button>
                    </form>
                </div>
                <div class="card">
                    <h2>History</h2>
                    {% if entries %}
                        <table>
                            <thead>
                                <tr>
                                    <th>Date</th>
                                    <th>Weight (lbs)</th>
                                    <th>Body Fat %</th>
                                    <th>Waist (in)</th>
                                    <th>Actions</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for entry in entries %}
                                    <tr>
                                        <td>{{ entry.date }}</td>
                                        <td>{{ '%.2f'|format(entry.weight) }}</td>
                                        <td>{{ '%.2f'|format(entry.body_fat) if entry.body_fat is not none else '–' }}</td>
                                        <td>{{ '%.2f'|format(entry.waist_size) if entry.waist_size is not none else '–' }}</td>
                                        <td>
                                            <a href="#" class="btn btn-secondary btn-sm" onclick="openEditModal('{{ entry.row_num }}', '{{ entry.date }}', '{{ entry.weight }}', '{{ entry.body_fat or '' }}', '{{ entry.waist_size or '' }}')">Edit</a>
                                            <a href="{{ url_for('delete_entry', row_index=entry.row_num) }}" class="btn btn-danger btn-sm">Delete</a>
                                        </td>
                                    </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    {% else %}
                        <p>No entries yet. Add one above to get started!</p>
                    {% endif %}
                </div>
            </div>
            <div class="right-column">
                <div class="card chart-card">
                    <h2>Weight Progress</h2>
                    {% if combined_labels %}
                        <canvas id="weightChart"></canvas>
                    {% else %}
                        <p>No data to display. Add a weight entry to see the chart.</p>
                    {% endif %}
                </div>
                <div class="card">
                    <h2>Summary</h2>
                    {% if summary_data %}
                        <table class="summary-table">
                            <tbody>
                                <tr>
                                    <td>Current Weight</td>
                                    <td>{{ '%.2f lbs'|format(summary_data.current) if summary_data.current is not none else 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td>Start Weight</td>
                                    <td>{{ '%.2f lbs'|format(summary_data.start) if summary_data.start is not none else 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td>Goal Weight</td>
                                    <td>{{ '%.2f lbs'|format(summary_data.goal) if summary_data.goal is not none else 'N/A' }}</td>
                                </tr>
                                <tr class="{{ summary_data.goal_class }}">
                                    <td>Weight to Goal</td>
                                    <td>{{ summary_data.to_goal if summary_data.to_goal is not none else 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td>Highest Weight</td>
                                    <td>{{ '%.2f lbs'|format(summary_data.highest) if summary_data.highest is not none else 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td>Lowest Weight</td>
                                    <td>{{ '%.2f lbs'|format(summary_data.lowest) if summary_data.lowest is not none else 'N/A' }}</td>
                                </tr>
                                {% if 'current_bf' in summary_data %}
                                <tr class="summary-divider"><td colspan="2"></td></tr>
                                <tr>
                                    <td>Current Body Fat</td>
                                    <td>{{ '%.2f%%'|format(summary_data.current_bf) }}</td>
                                </tr>
                                <tr>
                                    <td>Highest Body Fat</td>
                                    <td>{{ '%.2f%%'|format(summary_data.highest_bf) }}</td>
                                </tr>
                                <tr>
                                    <td>Lowest Body Fat</td>
                                    <td>{{ '%.2f%%'|format(summary_data.lowest_bf) }}</td>
                                </tr>
                                {% endif %}
                                {% if 'current_ws' in summary_data %}
                                <tr class="summary-divider"><td colspan="2"></td></tr>
                                <tr>
                                    <td>Current Waist Size</td>
                                    <td>{{ '%.2f in'|format(summary_data.current_ws) }}</td>
                                </tr>
                                <tr>
                                    <td>Highest Waist Size</td>
                                    <td>{{ '%.2f in'|format(summary_data.highest_ws) }}</td>
                                </tr>
                                <tr>
                                    <td>Lowest Waist Size</td>
                                    <td>{{ '%.2f in'|format(summary_data.lowest_ws) }}</td>
                                </tr>
                                {% endif %}
                            </tbody>
                        </table>
                    {% else %}
                        <p>No data for summary.</p>
                    {% endif %}
                </div>
            </div>
        </div>

        <div class="charts-grid">
            {% if body_fat_labels %}
            <div class="card">
                <h2>Body Fat % Progress</h2>
                <canvas id="bodyFatChart"></canvas>
            </div>
            {% endif %}

            {% if waist_size_labels %}
            <div class="card">
                <h2>Waist Size Progress (in)</h2>
                <canvas id="waistSizeChart"></canvas>
            </div>
            {% endif %}
        </div>

    </div>

    <div id="editModal" class="modal">
        <div class="modal-content">
            <span class="close-button" onclick="closeEditModal()">&times;</span>
            <h2>Edit Entry</h2>
            <form id="editForm" method="post">
                <div class="form-group">
                    <label for="edit_date">Date:</label>
                    <input type="date" id="edit_date" name="date" required>
                </div>
                <div class="form-group">
                    <label for="edit_weight">Weight (lbs):</label>
                    <input type="number" id="edit_weight" name="weight" step="0.01" required>
                </div>
                <div class="form-group">
                    <label for="edit_body_fat">Body Fat (%): <span class="optional-text">(Optional)</span></label>
                    <input type="number" id="edit_body_fat" name="body_fat" step="0.01">
                </div>
                <div class="form-group">
                    <label for="edit_waist_size">Waist Size (in): <span class="optional-text">(Optional)</span></label>
                    <input type="number" id="edit_waist_size" name="waist_size" step="0.01">
                </div>
                <button type="submit" class="btn btn-primary">Save Changes</button>
            </form>
        </div>
    </div>

    <script>
        // --- Main Weight Chart ---
        {% if combined_labels and chart_data_1 %}
        const ctx = document.getElementById('weightChart').getContext('2d');
        const chart_config = {{ chart_config | tojson }};
        const datasets = [{
            label: '{{ primary_user }} Weight (lbs)',
            data: {{ chart_data_1 | tojson }},
            borderColor: '#33CFFF',
            backgroundColor: 'rgba(51, 207, 255, 0.1)',
            yAxisID: 'y1',
            fill: true,
            tension: 0.1,
            spanGaps: true
        }];

        {% if comparison_user and chart_data_2_to_plot %}
        datasets.push({
            label: '{{ comparison_user }} Weight (lbs)',
            data: {{ chart_data_2_to_plot | tojson }},
            borderColor: '#9D63FF',
            backgroundColor: 'rgba(157, 99, 255, 0.1)',
            yAxisID: 'y2',
            fill: true,
            tension: 0.1,
            spanGaps: true
        });
        {% endif %}

        const chartData = { labels: {{ combined_labels | tojson }}, datasets: datasets };
        const chartOptions = {
            responsive: true,
            interaction: { mode: 'index', intersect: false },
            scales: { y1: { type: 'linear', display: true, position: 'left', title: { display: true, text: '{{ primary_user }} Weight (lbs)', color: '#33CFFF'}}},
            plugins: { tooltip: { callbacks: { label: function(context) { let label = context.dataset.label || ''; if (label) { label += ': '; } if (context.parsed.y !== null) { label += context.parsed.y.toFixed(2) + ' lbs'; } return label; }}}, annotation: { annotations: {}}}
        };

        if (chart_config.y2_axis_label) {
            chartOptions.scales.y2 = { type: 'linear', display: true, position: 'right', title: { display: true, text: chart_config.y2_axis_label, color: '#9D63FF' }, grid: { drawOnChartArea: false }};
        }
        {% if primary_user_data.start_weight %}
        chartOptions.plugins.annotation.annotations.startLine1 = { type: 'line', yMin: {{ primary_user_data.start_weight }}, yMax: {{ primary_user_data.start_weight }}, yScaleID: 'y1', borderColor: '#33CFFF', borderWidth: 2, borderDash: [6, 6], label: { content: 'Start: {{ "%.2f"|format(primary_user_data.start_weight) }} lbs', display: {{ 'false' if comparison_user else 'true' }}, position: 'start', backgroundColor: 'rgba(51, 207, 255, 0.8)' }};
        {% endif %}
        {% if primary_user_data.goal_weight %}
        chartOptions.plugins.annotation.annotations.goalLine1 = { type: 'line', yMin: {{ primary_user_data.goal_weight }}, yMax: {{ primary_user_data.goal_weight }}, yScaleID: 'y1', borderColor: 'var(--danger-color)', borderWidth: 2, borderDash: [6, 6], label: { content: 'Goal: {{ "%.2f"|format(primary_user_data.goal_weight) }} lbs', display: {{ 'false' if comparison_user else 'true' }}, position: 'end', backgroundColor: 'rgba(220, 53, 69, 0.8)' }};
        {% endif %}

        if (chart_config.y1_min !== null && chart_config.y1_max !== null) { chartOptions.scales.y1.min = chart_config.y1_min; chartOptions.scales.y1.max = chart_config.y1_max; }
        if (chart_config.y2_axis_label && chart_config.y2_min !== null && chart_config.y2_max !== null) { chartOptions.scales.y2.min = chart_config.y2_min; chartOptions.scales.y2.max = chart_config.y2_max; }
        new Chart(ctx, { type: 'line', data: chartData, options: chartOptions });
        {% endif %}

        // --- Body Fat Chart ---
        {% if body_fat_labels %}
        const bf_ctx = document.getElementById('bodyFatChart').getContext('2d');
        new Chart(bf_ctx, {
            type: 'line',
            data: {
                labels: {{ body_fat_labels | tojson }},
                datasets: [{
                    label: 'Body Fat (%)',
                    data: {{ body_fat_data | tojson }},
                    borderColor: '#28a745',
                    backgroundColor: 'rgba(40, 167, 69, 0.1)',
                    fill: true,
                    tension: 0.1,
                    spanGaps: true
                }]
            },
            options: { responsive: true, scales: { y: { title: { display: true, text: 'Body Fat (%)' }}}, plugins: { legend: { display: false }}}
        });
        {% endif %}

        // --- Waist Size Chart ---
        {% if waist_size_labels %}
        const ws_ctx = document.getElementById('waistSizeChart').getContext('2d');
        new Chart(ws_ctx, {
            type: 'line',
            data: {
                labels: {{ waist_size_labels | tojson }},
                datasets: [{
                    label: 'Waist Size (in)',
                    data: {{ waist_size_data | tojson }},
                    borderColor: '#fd7e14',
                    backgroundColor: 'rgba(253, 126, 20, 0.1)',
                    fill: true,
                    tension: 0.1,
                    spanGaps: true
                }]
            },
            options: { responsive: true, scales: { y: { title: { display: true, text: 'Waist Size (in)' }}}, plugins: { legend: { display: false }}}
        });
        {% endif %}

        // --- Modal Control Functions ---
        function openEditModal(row_index, date, weight, body_fat, waist_size) {
            const modal = document.getElementById('editModal');
            const form = document.getElementById('editForm');
            document.getElementById('edit_date').value = date;
            document.getElementById('edit_weight').value = weight;
            document.getElementById('edit_body_fat').value = body_fat;
            document.getElementById('edit_waist_size').value = waist_size;
            form.action = `/update/${row_index}`;
            modal.style.display = 'block';
        }

        function closeEditModal() {
            document.getElementById('editModal').style.display = 'none';
        }

        window.onclick = function(event) {
            const modal = document.getElementById('editModal');
            if (event.target == modal) {
                modal.style.display = 'none';
            }
        }
    </script>
</body>
</html>
"""

# --- CSS Content ---
CSS_CONTENT = """
:root {
    --primary-color: #33CFFF;
    --secondary-color: #9D63FF;
    --danger-color: #dc3545;
    --background-color: #f4f7f6;
    --card-background: #ffffff;
    --text-color: #333;
    --border-color: #e0e0e0;
    --shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
}
body {
    font-family: 'Inter', sans-serif;
    background-color: var(--background-color);
    color: var(--text-color);
    margin: 0;
    padding: 20px;
    display: flex;
    justify-content: center;
    align-items: flex-start;
    min-height: 100vh;
}
.container {
    width: 100%;
    max-width: 1200px;
    display: flex;
    flex-direction: column;
    gap: 20px;
}
h1, h2 {
    text-align: center;
}
h1 {
    color: var(--primary-color);
    margin-bottom: 10px;
}
.card {
    background-color: var(--card-background);
    border-radius: 12px;
    padding: 25px;
    box-shadow: var(--shadow);
    border: 1px solid var(--border-color);
}
.user-management-grid {
    display: flex;
    flex-direction: column;
    gap: 20px;
}
.user-selector-form {
    display: flex;
    flex-direction: column;
    gap: 15px;
}
.user-actions {
    display: flex;
    flex-direction: column;
    gap: 15px;
    margin-top: 10px;
    border-top: 1px solid var(--border-color);
    padding-top: 20px;
}
.user-actions .btn, .user-actions form:not(.add-user-form) .btn {
    width: 100%;
}
.input-with-button {
    display: flex;
    gap: 10px;
}
.input-with-button input {
    flex-grow: 1;
}
.goal-form {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    align-items: end;
}
.log-entry-form .btn-primary {
    width: 100%;
}
.log-entry-form-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    margin-bottom: 20px;
}
.log-entry-form-grid .form-group {
    margin-bottom: 0;
}
.form-group {
    margin-bottom: 20px;
}
label {
    display: block;
    margin-bottom: 8px;
    font-weight: 500;
}
.optional-text {
    font-weight: 400;
    color: #777;
    font-size: 0.9em;
}
input[type="number"], input[type="text"], input[type="date"], select {
    width: 100%;
    padding: 12px;
    border: 1px solid var(--border-color);
    border-radius: 8px;
    box-sizing: border-box;
    font-size: 1rem;
}
select { padding: 10px; background-color: white; }
.btn {
    padding: 12px 20px;
    border: none;
    border-radius: 8px;
    cursor: pointer;
    font-size: 1rem;
    font-weight: 500;
    transition: all 0.2s ease-in-out;
    text-decoration: none;
    display: inline-block;
    text-align: center;
}
.btn:hover { opacity: 0.9; }
.btn-primary { background-color: var(--primary-color); color: white; }
.btn-primary:hover { transform: translateY(-2px); }
.btn-secondary { background-color: var(--secondary-color); color: white; }
.btn-sm { padding: 5px 10px; font-size: 0.875rem; }
.goal-form .btn-secondary { width: 100%; }
.btn-danger { background-color: var(--danger-color); color: white; }
table {
    width: 100%;
    border-collapse: collapse;
    margin-top: 10px;
}
th, td {
    padding: 12px;
    text-align: left;
    border-bottom: 1px solid var(--border-color);
}
td { text-align: center; }
td:first-child { text-align: left; }
thead th { font-weight: 600; text-align: center; }
thead th:first-child { text-align: left; }
tbody tr:last-child td { border-bottom: none; }
.main-grid {
    display: grid;
    grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
    gap: 20px;
    align-items: start;
}
.left-column {
    display: flex;
    flex-direction: column;
    min-width: 0;
    gap: 20px;
}
.right-column {
    position: sticky;
    top: 20px;
    display: flex;
    flex-direction: column;
    gap: 20px;
}
.summary-table { margin-top: 0; }
.summary-table td { padding: 10px 0; }
.summary-table td:first-child { font-weight: 500; color: #555; text-align: left; }
.summary-table td:last-child { text-align: right; font-weight: 600; }
.summary-divider td { padding: 8px 0; border-bottom: 1px solid var(--border-color); }
.goal-positive { color: #155724; }
.goal-negative { color: #721c24; }
.charts-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
    gap: 20px;
}
.modal {
    display: none; position: fixed; z-index: 1000; left: 0; top: 0;
    width: 100%; height: 100%; overflow: auto; background-color: rgba(0,0,0,0.5);
}
.modal-content {
    background-color: #fefefe; margin: 10% auto; padding: 30px;
    border: 1px solid #888; width: 80%; max-width: 500px;
    border-radius: 12px; position: relative;
}
.close-button {
    color: #aaa; position: absolute; top: 10px; right: 20px;
    font-size: 28px; font-weight: bold; cursor: pointer;
}
.close-button:hover, .close-button:focus { color: black; text-decoration: none; }
.flash-messages { list-style: none; padding: 0; margin-bottom: 1rem; }
.flash { padding: 1rem; border-radius: 8px; margin-bottom: 0.5rem; text-align: center; }
.flash.success { background-color: #d4edda; color: #155724; }
.flash.error { background-color: #f8d7da; color: #721c24; }
"""

# 1. Setup: Create directories and files
def setup_environment():
    """Creates and validates necessary directories and files."""
    print("Validating application environment...")
    os.makedirs(STATIC_FOLDER, exist_ok=True)

    with open(CSS_FILE, 'w') as f:
        f.write(CSS_CONTENT)
    print(f"Ensured '{CSS_FILE}' is up to date.")

    DATA_HEADERS = ["Date", "Weight (lbs)", "User", "Body Fat %", "Waist Size (in)"]
    USER_HEADERS = ["Username", "Start Weight (lbs)", "Goal Weight (lbs)"]

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet_data = workbook.active
        sheet_data.title = "Weight Data"
        sheet_data.append(DATA_HEADERS)
        sheet_users = workbook.create_sheet("Users")
        sheet_users.append(USER_HEADERS)
        sheet_users.append(["User 1", None, None])
        workbook.save(EXCEL_FILE)
        print(f"Created '{EXCEL_FILE}' with required sheets and headers.")
    else:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        updated = False
        if "Users" not in workbook.sheetnames:
            sheet_users = workbook.create_sheet("Users")
            sheet_users.append(USER_HEADERS)
            updated = True
            print("Added missing 'Users' sheet.")
        if "Weight Data" not in workbook.sheetnames:
            sheet_data = workbook.create_sheet("Weight Data")
            sheet_data.append(DATA_HEADERS)
            updated = True
            print("Added missing 'Weight Data' sheet.")
        else:
            sheet_data = workbook["Weight Data"]
            current_headers = [cell.value for cell in sheet_data[1]]
            if "Body Fat %" not in current_headers:
                sheet_data.cell(row=1, column=sheet_data.max_column + 1, value="Body Fat %")
                updated = True
            if "Waist Size (in)" not in current_headers:
                sheet_data.cell(row=1, column=sheet_data.max_column + 1, value="Waist Size (in)")
                updated = True
            if updated: print("Updated headers in 'Weight Data' sheet.")
        if updated: workbook.save(EXCEL_FILE)

setup_environment()
app = Flask(__name__, static_folder=STATIC_FOLDER)
app.secret_key = 'a_secure_random_secret_key'

def get_users():
    """Reads the list of users from the 'Users' sheet."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Users"]
        users = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]
        return users
    except (FileNotFoundError, KeyError):
        return ["User 1"]

def get_user_data(user):
    """Reads start and goal weight for a specific user."""
    raw_data = {"start_weight": None, "goal_weight": None}
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Users"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user:
                raw_data = {"start_weight": row[1], "goal_weight": row[2]}
                break
    except (FileNotFoundError, KeyError):
        return raw_data
    start_weight, goal_weight = None, None
    try:
        if raw_data["start_weight"] is not None: start_weight = float(raw_data["start_weight"])
    except (ValueError, TypeError): pass
    try:
        if raw_data["goal_weight"] is not None: goal_weight = float(raw_data["goal_weight"])
    except (ValueError, TypeError): pass
    return {"start_weight": start_weight, "goal_weight": goal_weight}

def get_weight_entries(active_user):
    """Reads all entries for a specific user from the Excel file."""
    entries = []
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Weight Data"]
        for index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row_values) >= 3 and row_values[2] == active_user:
                date_val, weight_val = row_values[0], row_values[1]
                body_fat_val = row_values[3] if len(row_values) > 3 else None
                waist_size_val = row_values[4] if len(row_values) > 4 else None
                if date_val is None or weight_val is None: continue
                try:
                    numeric_weight = float(weight_val)
                    numeric_body_fat = float(body_fat_val) if body_fat_val is not None else None
                    numeric_waist_size = float(waist_size_val) if waist_size_val is not None else None
                except (ValueError, TypeError): continue
                normalized_date = (date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime.datetime) else str(date_val).split(' ')[0])
                entries.append({"date": normalized_date, "weight": numeric_weight, "body_fat": numeric_body_fat, "waist_size": numeric_waist_size, "row_num": index})
    except FileNotFoundError: return []
    return sorted(entries, key=lambda x: x['date'], reverse=True)

def add_weight_entry(date_str, weight, user, body_fat, waist_size):
    """Adds a new entry to the Excel file."""
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Weight Data"]
    sheet.append([date_str, weight, user, body_fat, waist_size])
    workbook.save(EXCEL_FILE)

def update_weight_entry(row_index, new_date, new_weight, new_body_fat, new_waist_size):
    """Updates an existing entry by its row index."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Weight Data"]
        if 1 < row_index <= sheet.max_row:
            sheet.cell(row=row_index, column=1).value = new_date
            sheet.cell(row=row_index, column=2).value = new_weight
            sheet.cell(row=row_index, column=4).value = new_body_fat
            sheet.cell(row=row_index, column=5).value = new_waist_size
            workbook.save(EXCEL_FILE)
            return True
    except Exception: return False
    return False

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        user, date_str = request.form.get('user'), request.form.get('date')
        try:
            weight = float(request.form['weight'])
            body_fat = float(s) if (s := request.form.get('body_fat')) else None
            waist_size = float(s) if (s := request.form.get('waist_size')) else None
            if weight > 0 and date_str:
                add_weight_entry(date_str, weight, user, body_fat, waist_size)
                flash('Entry added successfully!', 'success')
            else:
                flash('Weight must be a positive number.', 'error')
        except (ValueError, TypeError):
            flash('Invalid input. Please enter valid numbers.', 'error')
        return redirect(url_for('index', user1=user))

    all_users = get_users()
    if not all_users:
        return render_template_string(HTML_CONTENT, all_users=[], primary_user=None, entries=[], primary_user_data={})

    primary_user = request.args.get('user1', all_users[0])
    comparison_user = request.args.get('user2')
    entries = get_weight_entries(primary_user)
    primary_user_data = get_user_data(primary_user)
    comparison_user_data = get_user_data(comparison_user) if comparison_user else {}
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # --- Weight Chart Data ---
    entries1 = sorted([e for e in entries if e.get('weight') is not None], key=lambda x: x['date'])
    data1_map = {e['date']: e['weight'] for e in entries1}
    y1_values = list(data1_map.values())
    if primary_user_data.get('start_weight') is not None: y1_values.append(primary_user_data['start_weight'])
    if primary_user_data.get('goal_weight') is not None: y1_values.append(primary_user_data['goal_weight'])
    y1_min, y1_max = (None, None)
    if y1_values:
        padding = 5
        y1_min, y1_max = min(y1_values) - padding, max(y1_values) + padding
    chart_config = {"y1_min": y1_min, "y1_max": y1_max, "y2_axis_label": None, "y2_min": None, "y2_max": None}
    combined_labels, chart_data_1, chart_data_2_to_plot = [], [], []

    if comparison_user:
        entries2 = sorted(get_weight_entries(comparison_user), key=lambda x: x['date'])
        data2_map = {e['date']: e['weight'] for e in entries2}
        combined_labels = sorted(list(set(data1_map.keys()) | set(data2_map.keys())))
        chart_data_1 = [data1_map.get(date) for date in combined_labels]
        chart_data_2_to_plot = [data2_map.get(date) for date in combined_labels]
        chart_config["y2_axis_label"] = f"{comparison_user} Weight (lbs)"

        # --- MODIFIED: Y-Axis Scaling Logic ---
        p_start = primary_user_data.get('start_weight')
        p_goal = primary_user_data.get('goal_weight')
        c_start = comparison_user_data.get('start_weight')
        c_goal = comparison_user_data.get('goal_weight')

        # Check if all required values are present and valid for proportional scaling
        if (y1_max is not None and y1_min is not None and
            p_start and p_start > 0 and p_goal and p_goal > 0 and
            c_start and c_goal):
            
            # Proportional scaling based on start and goal weights
            chart_config['y2_max'] = y1_max * (c_start / p_start)
            chart_config['y2_min'] = y1_min * (c_goal / p_goal)
        else:
            # Fallback to default auto-scaling if goal/start weights are missing
            y2_values = [v for v in chart_data_2_to_plot if v is not None]
            if c_start is not None: y2_values.append(c_start)
            if c_goal is not None: y2_values.append(c_goal)
            
            if y2_values:
                padding = 5
                chart_config['y2_min'], chart_config['y2_max'] = min(y2_values) - padding, max(y2_values) + padding
    else:
        combined_labels, chart_data_1 = [e['date'] for e in entries1], [e['weight'] for e in entries1]


    # --- Summary Data ---
    summary_data = {}
    if entries:
        all_weights = [e['weight'] for e in entries]
        summary_data.update({'current': all_weights[0], 'highest': max(all_weights), 'lowest': min(all_weights)})
    summary_data.update({'start': primary_user_data.get('start_weight'), 'goal': primary_user_data.get('goal_weight')})
    if summary_data.get('current') and summary_data.get('goal'):
        to_goal = summary_data['current'] - summary_data['goal']
        if to_goal > 0.05: summary_data.update({'to_goal': f"{to_goal:.2f} lbs to lose", 'goal_class': 'goal-negative'})
        elif to_goal < -0.05: summary_data.update({'to_goal': f"{-to_goal:.2f} lbs below goal", 'goal_class': 'goal-positive'})
        else: summary_data.update({'to_goal': "Goal reached!", 'goal_class': 'goal-positive'})

    # --- Optional Charts & Summary Stats ---
    body_fat_entries = [e for e in entries if e.get('body_fat') is not None]
    body_fat_labels, body_fat_data = [], []
    if body_fat_entries:
        sorted_bf = sorted(body_fat_entries, key=lambda x: x['date'])
        body_fat_labels, body_fat_data = [e['date'] for e in sorted_bf], [e['body_fat'] for e in sorted_bf]
        all_bf = [e['body_fat'] for e in body_fat_entries]
        summary_data.update({'current_bf': body_fat_entries[0]['body_fat'], 'highest_bf': max(all_bf), 'lowest_bf': min(all_bf)})

    waist_size_entries = [e for e in entries if e.get('waist_size') is not None]
    waist_size_labels, waist_size_data = [], []
    if waist_size_entries:
        sorted_ws = sorted(waist_size_entries, key=lambda x: x['date'])
        waist_size_labels, waist_size_data = [e['date'] for e in sorted_ws], [e['waist_size'] for e in sorted_ws]
        all_ws = [e['waist_size'] for e in waist_size_entries]
        summary_data.update({'current_ws': waist_size_entries[0]['waist_size'], 'highest_ws': max(all_ws), 'lowest_ws': min(all_ws)})

    return render_template_string(
        HTML_CONTENT, entries=entries, primary_user=primary_user, comparison_user=comparison_user,
        all_users=all_users, primary_user_data=primary_user_data, comparison_user_data=comparison_user_data,
        combined_labels=combined_labels, chart_data_1=chart_data_1, chart_data_2_to_plot=chart_data_2_to_plot,
        chart_config=chart_config, today_date=today_date, summary_data=summary_data,
        body_fat_labels=body_fat_labels, body_fat_data=body_fat_data,
        waist_size_labels=waist_size_labels, waist_size_data=waist_size_data
    )

@app.route('/update_goals', methods=['POST'])
def update_goals():
    user = request.form.get('user')
    if not user:
        flash("No user selected.", "error")
        return redirect(url_for('index'))
    try:
        start_weight_val = float(s) if (s := request.form.get('start_weight')) else None
        goal_weight_val = float(s) if (s := request.form.get('goal_weight')) else None
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Users"]
        user_found = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == user:
                row[1].value, row[2].value = start_weight_val, goal_weight_val
                user_found = True
                break
        if user_found:
            workbook.save(EXCEL_FILE)
            flash(f"Goals for {user} updated successfully!", "success")
        else:
            flash(f"Could not find user {user} to update.", "error")
    except ValueError: flash("Invalid input for weights.", "error")
    except Exception as e: flash(f"An error occurred: {e}", "error")
    return redirect(url_for('index', user1=user))

@app.route('/add_user', methods=['POST'])
def add_user():
    new_user_name = request.form.get('new_user_name', '').strip()
    if not new_user_name:
        flash("User name cannot be empty.", "error")
        return redirect(url_for('index'))
    if new_user_name in get_users():
        flash(f"User '{new_user_name}' already exists.", "error")
        return redirect(url_for('index'))
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Users"]
    sheet.append([new_user_name, None, None])
    workbook.save(EXCEL_FILE)
    flash(f"User '{new_user_name}' added successfully!", "success")
    return redirect(url_for('index', user1=new_user_name))

@app.route('/update/<int:row_index>', methods=['POST'])
def update_entry(row_index):
    new_date = request.form.get('date')
    try:
        new_weight = float(request.form.get('weight'))
        new_body_fat = float(s) if (s := request.form.get('body_fat')) else None
        new_waist_size = float(s) if (s := request.form.get('waist_size')) else None
        if new_weight > 0 and new_date:
            if update_weight_entry(row_index, new_date, new_weight, new_body_fat, new_waist_size):
                flash('Entry updated successfully!', 'success')
            else:
                flash('Could not find entry to update.', 'error')
        else:
            flash('Invalid weight or date.', 'error')
    except (ValueError, TypeError):
        flash('Invalid input for weight.', 'error')
    return redirect(request.referrer or url_for('index'))

@app.route('/delete/<int:row_index>')
def delete_entry(row_index):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook["Weight Data"]
        if 1 < row_index <= sheet.max_row:
            sheet.delete_rows(row_index)
            workbook.save(EXCEL_FILE)
            flash('Entry deleted successfully!', 'success')
        else:
            flash('Could not find the entry to delete.', 'error')
    except Exception as e:
        flash(f'An error occurred: {e}', 'error')
    return redirect(request.referrer or url_for('index'))

@app.route('/delete_user', methods=['POST'])
def delete_user():
    user_to_delete = request.form.get('user')
    if len(get_users()) <= 1:
        flash("Cannot delete the last user.", "error")
        return redirect(url_for('index', user1=user_to_delete))
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        for sheet_name in ["Weight Data", "Users"]:
            sheet = workbook[sheet_name]
            col_idx = 3 if sheet_name == "Weight Data" else 1
            rows_to_delete = [r.row for r in sheet.iter_rows(min_row=2) if r[col_idx - 1].value == user_to_delete]
            for r_idx in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(r_idx)
        workbook.save(EXCEL_FILE)
        flash(f"User '{user_to_delete}' and all data have been deleted.", "success")
    except Exception as e:
        flash(f"An error occurred: {e}", "error")
        return redirect(url_for('index', user1=user_to_delete))
    return redirect(url_for('index'))

if __name__ == '__main__':
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        print("\n--- Starting Flask Server ---")
        print("Open your web browser and go to: http://127.0.0.1:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)


